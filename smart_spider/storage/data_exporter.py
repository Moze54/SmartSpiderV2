"""
数据导出模块 - 基础版本
"""
import json
import csv
import os
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import logging

from smart_spider.core.exceptions import StorageException
from smart_spider.core.logger import logger


class DataExporter:
    """数据导出器"""

    def __init__(self):
        self.output_dir = Path("output")
        self.output_dir.mkdir(exist_ok=True)
        self.export_stats = {
            "total_exports": 0,
            "total_files": 0,
            "total_records": 0,
            "formats_used": set(),
            "last_export": None
        }

    def auto_export(self, data: List[Dict[str, Any]], filename: str, task_id: str, formats: List[str]) -> Dict[str, str]:
        """自动导出数据到多种格式"""
        exported_files = {}

        for format_type in formats:
            try:
                if format_type == "json":
                    file_path = self.export_to_json(data, filename)
                elif format_type == "csv":
                    file_path = self.export_to_csv(data, filename)
                elif format_type == "excel":
                    file_path = self.export_to_excel(data, filename)
                else:
                    logger.warning(f"不支持的导出格式: {format_type}")
                    continue

                exported_files[format_type] = file_path
                self.export_stats["formats_used"].add(format_type)
                logger.info(f"数据导出成功: {format_type} -> {file_path}")

            except Exception as e:
                logger.error(f"数据导出失败: {format_type}, 错误: {str(e)}")

        # 更新统计信息
        self.export_stats["total_exports"] += 1
        self.export_stats["total_files"] += len(exported_files)
        self.export_stats["total_records"] += len(data)
        self.export_stats["last_export"] = datetime.now().isoformat()

        return exported_files

    def export_to_json(self, data: List[Dict[str, Any]], filename: str) -> str:
        """导出到JSON格式"""
        try:
            file_path = self.output_dir / f"{filename}.json"

            export_data = {
                "metadata": {
                    "exported_at": datetime.now().isoformat(),
                    "record_count": len(data),
                    "format": "json"
                },
                "data": data
            }

            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)

            return str(file_path)

        except Exception as e:
            raise StorageException(f"JSON导出失败: {str(e)}", storage_type="json")

    def export_to_csv(self, data: List[Dict[str, Any]], filename: str) -> str:
        """导出到CSV格式"""
        try:
            if not data:
                raise StorageException("没有数据可导出", storage_type="csv")

            file_path = self.output_dir / f"{filename}.csv"

            # 获取所有可能的字段名
            fieldnames = set()
            for item in data:
                fieldnames.update(item.keys())
            fieldnames = sorted(list(fieldnames))

            with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(data)

            return str(file_path)

        except Exception as e:
            raise StorageException(f"CSV导出失败: {str(e)}", storage_type="csv")

    def export_to_excel(self, data: List[Dict[str, Any]], filename: str) -> str:
        """导出到Excel格式"""
        try:
            if not data:
                raise StorageException("没有数据可导出", storage_type="excel")

            file_path = self.output_dir / f"{filename}.xlsx"

            # 尝试使用openpyxl，如果没有安装则回退到CSV
            try:
                import openpyxl
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter

                wb = Workbook()
                ws = wb.active

                # 获取所有可能的字段名
                fieldnames = set()
                for item in data:
                    fieldnames.update(item.keys())
                fieldnames = sorted(list(fieldnames))

                # 写入表头
                for col, fieldname in enumerate(fieldnames, 1):
                    ws.cell(row=1, column=col, value=fieldname)

                # 写入数据
                for row, item in enumerate(data, 2):
                    for col, fieldname in enumerate(fieldnames, 1):
                        value = item.get(fieldname, "")
                        ws.cell(row=row, column=col, value=str(value))

                # 自动调整列宽
                for col in range(1, len(fieldnames) + 1):
                    ws.column_dimensions[get_column_letter(col)].auto_size = True

                wb.save(file_path)
                return str(file_path)

            except ImportError:
                logger.warning("openpyxl未安装，回退到CSV格式")
                return self.export_to_csv(data, filename)

        except Exception as e:
            raise StorageException(f"Excel导出失败: {str(e)}", storage_type="excel")

    def get_export_stats(self) -> Dict[str, Any]:
        """获取导出统计信息"""
        return {
            "total_exports": self.export_stats["total_exports"],
            "total_files": self.export_stats["total_files"],
            "total_records": self.export_stats["total_records"],
            "formats_used": list(self.export_stats["formats_used"]),
            "last_export": self.export_stats["last_export"]
        }

    def cleanup_old_exports(self, days_to_keep: int = 7) -> int:
        """清理旧的导出文件"""
        removed_count = 0
        cutoff_time = datetime.now() - timedelta(days=days_to_keep)

        try:
            for file_path in self.output_dir.glob("*"):
                if file_path.is_file():
                    file_mtime = datetime.fromtimestamp(file_path.stat().st_mtime)
                    if file_mtime < cutoff_time:
                        file_path.unlink()
                        removed_count += 1
                        logger.info(f"删除旧导出文件: {file_path}")

            logger.info(f"清理完成: 删除了 {removed_count} 个旧文件")
            return removed_count

        except Exception as e:
            logger.error(f"清理旧导出文件失败: {str(e)}")
            return 0


# 全局数据导出器实例
data_exporter = DataExporter()


# 快捷函数
def get_export_stats() -> Dict[str, Any]:
    """获取导出统计信息"""
    return data_exporter.get_export_stats()


def export_task_results(task_id: str, results: List[Dict[str, Any]], filename: Optional[str] = None, formats: List[str] = None) -> Dict[str, str]:
    """导出任务结果"""
    if not filename:
        filename = f"task_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    if not formats:
        formats = ['json', 'csv', 'excel']

    return data_exporter.auto_export(results, filename, task_id, formats)


__all__ = [
    'DataExporter',
    'data_exporter',
    'get_export_stats',
    'export_task_results'
]